#aula 5 abrindo arquivos excel

from openpyxl import load_workbook

# Abrir o arquivo e a planilha
wb = load_workbook("Teste.xlsx", data_only=True)
p= wb["Planilha1"]

# Ler as 5 primeiras entradas da 2ª coluna (coluna D)
valores = []
for i in range(1, 7):   # linhas 1 até 6
    celula = p.cell(row=i, column=2)   # column=2 significa coluna B
    valores.append(celula.value)

print(valores)


x=p['A']  # coluna A do Excel
y=p['B']  # coluna B


print([c.value for c in x])
print([c.value for c in y])

lin=p.max_row   # número de linhas
col=p.max_column   # número de colunas

# Entradas específicas da planilha (lembrando que as células são contadas a partir de 0)

# imprime célula B2 (linha 2, coluna 2) da planilha
# imprime célula A5 (linha 4, coluna 1) da planilha
 
print(p['B2'].value)
print(p['A5'].value)


lista=[]

for i in range(1,7):
    lista.append(p.cell(row=i, column=1).value)
print(lista)


# Mas é muito mais fácil fazer:

for cell in p['A']:
    print(cell.value)