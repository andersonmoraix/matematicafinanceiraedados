#aula 7
# APLICAÇÃO 2: Boxplot com dados da empresa XYZW3

# VARIAÇÕES AO LONGO DO ANO - voltamos a uma lista para construir
# um boxplot *por dia*

from openpyxl import load_workbook
import matplotlib.pyplot as fig
import numpy as np


wb=load_workbook("XYZW3.xlsx", data_only=True)
p= wb["Planilha1"]

lin=p.max_row

op=np.zeros(lin)
hig=np.zeros(lin)
low=np.zeros(lin)
fec=np.zeros(lin)

for i in range(1,lin):    
 op[i]=p.cell(row=i,column=2).value
 hig[i]=p.cell(row=i,column=3).value
 low[i]=p.cell(row=i,column=4).value
 fec[i]=p.cell(row=i,column=5).value

dados=[]

for i in range(lin):
 dados.append([op[i],hig[i],low[i],fec[i]])

fig.boxplot(dados)   # 1 ano de dados de boxplot